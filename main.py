import requests  
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os
from dotenv import load_dotenv
import time
import re

print("##############################################")
print("###   GERADOR DE FICHAS DE IMPLANTAÇÃO     ###")
print("###   Autor: Mauricio Luan /2025           ###")
print("##############################################\n")

#Esse laço while é para que o programa continue rodando até que o usuário decida encerrar
while True:
    #essa função vai pedir ao usuario  duas informações: o id do cliente e o id do chamado. Com essas informações,
    #o programa vai realizar as requisições na api do tomticket através de dois endpoints específicos. (Consta na 
    #documentação da api do tomticket). Os endpoints e o token de acesso estão salvos no arquivo.env. Esse arquivo
    #eu nao vou subir para o github pois são privados.
    def request():
        load_dotenv()
        API_TOKEN = os.getenv("API_TOKEN")
        API_URL = os.getenv("API_URL")
        API_TICKET_URL = os.getenv("API_TICKET_URL")
 
        id = input("\nID do cliente: ")  
        ticket_id = input("ID do chamado: ")

        api = f"{API_URL}{id}" 
        apiticket = f"{API_TICKET_URL}{ticket_id}"

        headers = {'Authorization': f'Bearer {API_TOKEN}'}

        resposta = requests.get(api, headers=headers) 
        resposta_ticket = requests.get(apiticket, headers=headers)

        return resposta, resposta_ticket
    resposta, resposta_ticket = request()


    #esse bloco cria um dicionario chamado arquivo, depois entra num laço que verifica o retorno das requisições.
    #Se der tudo certo ele cria um json para cada uma e armazena no dicionario, noemando uma chave para cada json.
    #Se o retorno for qualquer coisa diferente de 200 (ok), entra no else, joga a mensagem de retorno na tela e chama
    #a funcao request() novamente.
    arquivo = {}
    while True:
        if resposta.status_code == 200 and resposta_ticket.status_code == 200:
            arquivo["resposta"] = resposta.json() 
            arquivo["resposta_ticket"] = resposta_ticket.json()
            break
        else:
            print(f"Retorno da requisição do cliente: {resposta.status_code} - {resposta.reason}")
            print(f"Retorno da requisição do chamado: {resposta_ticket.status_code} - {resposta_ticket.reason}")
            resposta, resposta_ticket = request()


    #esse primeiro bloco vai pegar os dois json que estão dentro do dicionario 'arquivo' vai criar um arquivo chamado
    #'arquivo.json' e jogar o conteudo de cada json dentro dele. Aí ele cria a variavel file e joga o arquivo.json pra
    #dentro de dela. Na liinha de baixo o arquivo.json é lido e jogado dentro de um dicionario chamado dicionario. Isso
    #é feito pra poder acessar as chaves do json de forma mais fácil em um unico lugar.
    with open('arquivo.json', 'w', encoding='utf-8') as file:
        json.dump(arquivo, file, ensure_ascii=False, indent=4)

    with open('arquivo.json', 'r', encoding='utf-8') as file:
        dicionario = json.load(file)

    
    #aqui eu crio um dicionario vazio chamado dados_filtrados. Embaixo eu crio chaves dentro desse dicionario nomeadas
    #com os campos que vou precisar colocar na planilha depois. Em cada linha eu tento acessar os valores das chaves que
    #foram armazenados em 'dicionario'. Por ex: 'Chamado' vai receber o valor da chave 'protocol' que está dentro de 'data',
    #que está dentro de 'resposta_ticket' que está dentro de 'dicionario'.
    dados_filtrados = {}
    try:
        dados_filtrados['Chamado'] = dicionario['resposta_ticket']['data']['protocol']
        dados_filtrados['Razão Social'] = dicionario['resposta']['data'][0]['name'].strip()
        nome_fantasia = dicionario['resposta']['data'][0]['custom_fields'][0]['value']

        conta_empresa_loja = dicionario['resposta']['data'][0]['internal_id']
        conta, empresa, loja = conta_empresa_loja.split('-')
        dados_filtrados['Conta'] = f"{conta} - {dados_filtrados['Razão Social']}"
        dados_filtrados['Empresa'] = f"{empresa} - {dados_filtrados['Razão Social']}"
        dados_filtrados['Loja'] = f"{loja} - {nome_fantasia}"
        partes = conta_empresa_loja.split('-')
        conta_editada = partes[1] + partes[2]
        dados_filtrados['pdv'] = dicionario['resposta_ticket']['data']['subject']
        padrao = r"Terminais \[(\d+)\]" #Isso aqui só descobri catando na internet. É muito específico...
        resultado = re.search(padrao, dados_filtrados['pdv'])

        #essa parte foi gambiarra pra eu concatenar os valores de empresa, loja e a quantidade de terminais que o cliente tem.
        #Não consegui fazer sozinho e corri para o chatgpt kkkkkkkk.
        if resultado:
            numero_terminais = int(resultado.group(1))
            partes = conta_empresa_loja.split('-')
            conta_editada = partes[1] + partes[2]
            sequencias = [f"{conta_editada}{i:02}" for i in range(1, numero_terminais + 1)]

        #esses dados são os que estão dentro de 'custom_fields' que é um array de objetos. Eu criei uma lista chamada
        #'dados_a_capturar'        com os camposs que eu quero pegar. Depois eu crio um laço que vai percorrer cada item
        #do arraye verificar se o nome do campo é igual a algum campo da lista 'dados_a_capturar'. Se for igual, ele vai
        #pegar o valor do campo e armazenar no dicionario 'dados_filtrados'.
        captura = dicionario['resposta']['data'][0]['custom_fields']
        dados_a_capturar = [
            "Nome Fantasia","CNPJ","Endereco","Numero","Bairro","Cidade","COMERCIAL - Contato","COMERCIAL - Telefone","COMERCIAL - E-mail"
        ]

        for campo in dados_a_capturar:
            for item in captura:
                if item['name'] == campo:
                    dados_filtrados[campo] = item['value']

    except KeyError as e:
        print(f"Erro ao acessar uma chave do JSON: {e}")
    except IndexError as e:
        print(f"Erro ao acessar um índice: {e}")


    #nessa parte o que eu faço é só organizar os dados filtrados do json para que fiquem na ordem em que preciso.
    #Jogo essa organização dentro de dados_filtrados_reorganizado. O modelo de ficha que eu faço manualmente segue
    #exatamente essa ordem. Tem o detalhe de conter a data do lado do numero do chamado, por isso a funcao datetime.date.today().
    data = datetime.date.today()
    dados_filtrados_reorganizado = {
        "Chamado": str(dados_filtrados["Chamado"]) + " - " + str(data.strftime("%d/%m/%Y")),
        "Nome Fantasia": dados_filtrados["Nome Fantasia"],
        "Razão Social": dados_filtrados["Razão Social"],
        "CNPJ": dados_filtrados["CNPJ"],
        #"Endereco": dados_filtrados["Endereco"] + ", " + dados_filtrados["Numero"],
        "Endereco": str(dados_filtrados["Endereco"] or "") + ", " + str(dados_filtrados["Numero"] or ""),
        "Bairro": dados_filtrados["Bairro"],
        "Cidade": dados_filtrados["Cidade"],
        "Contato": dados_filtrados["COMERCIAL - Contato"],
        "Telefone": dados_filtrados["COMERCIAL - Telefone"],
        "E-mail": dados_filtrados["COMERCIAL - E-mail"],
        "Conta": dados_filtrados["Conta"],
        "Empresa": dados_filtrados["Empresa"],
        "Loja": dados_filtrados["Loja"],
        "Token Payer": " / ".join(sequencias)
    }

    #feito a reorganização na parte de cima, eu crio um arquivo chamado 'teste' e jogo essa nova filtragem pra dentro desse arquivo.
    with open('teste', 'w', encoding='utf-8') as file:
        json.dump(dados_filtrados_reorganizado, file, ensure_ascii=False, indent=4) 
        

    #essa função é a que cria a planilha excel. Não entendo nada dessa parte, só copiei e colei do gpt kkkkkkkkkkkkk.
    #Mas assim, eu queria manter o mesmo estilo da planilha que eu fazia. Então eu printei a planilha e joguei no gpt,
    #pedindo para ele me dar um código base que me desse um resultado muito parecido. Depois disso eu só fui alterando
    #alguns detalhes até chegar no resultado que eu queria. A lib que eu usei pra fazer essa manipulação de é a openpyxl.
    def gerar_planilha_estilizada(dados, arquivo_excel, caminho_imagem):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sitef"

        fonte_negrito = Font(bold=True)
        fonte_branca = Font(color="FFFFFF", bold=True)
        fonte_cinza_claro = Font(color="808080", bold=True)
        alinhamento_central = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)
        borda_fina = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        preenchimento_cabecalho = PatternFill(start_color="171717", end_color="171717", fill_type="solid")
        preenchimento_secoes = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['B'].width = 50

        ws["A1"].value = "PAYER"
        ws["A1"].font = fonte_branca
        ws["A1"].alignment = alinhamento_central
        ws["A1"].fill = preenchimento_cabecalho
        ws["A1"].border = borda_fina

        img = Image(caminho_imagem)
        img.width = 120  
        img.height = 19  
        ws.add_image(img, 'A1') 

        ws["B1"].value = "FICHA DE IMPLANTAÇÃO"
        ws["B1"].font = fonte_branca
        ws["B1"].alignment = alinhamento_central
        ws["B1"].fill = preenchimento_cabecalho
        ws["B1"].border = borda_fina

        linha_atual = 2
        for chave, valor in dados.items():
            if chave == "Conta":
                ws.merge_cells(f"A{linha_atual}:B{linha_atual}")
                ws[f"A{linha_atual}"].value = "DADOS PAYER"
                ws[f"A{linha_atual}"].font = fonte_negrito
                ws[f"A{linha_atual}"].alignment = alinhamento_central
                ws[f"A{linha_atual}"].fill = preenchimento_secoes
                ws[f"A{linha_atual}"].border = borda_fina
                linha_atual += 1

            ws[f"A{linha_atual}"].value = chave.upper()
            ws[f"A{linha_atual}"].font = fonte_cinza_claro
            ws[f"A{linha_atual}"].border = borda_fina
            ws[f"A{linha_atual}"].alignment = alinhamento_central

            
            ws[f"B{linha_atual}"].value = valor
            ws[f"B{linha_atual}"].border = borda_fina
            ws[f"B{linha_atual}"].alignment = alinhamento_esquerda
            ws.column_dimensions['B'].width = 49
            
            if linha_atual in [13, 14, 15, 16]:
                ws[f"B{linha_atual}"].alignment = alinhamento_central
                ws.column_dimensions['B'].width = 48

            linha_atual += 1

        wb.save(arquivo_excel)
        print(f"\nPlanilha salva em: {arquivo_excel}")

    #aqui é como eu salvo a planilha. Primeiro eu defini o caminho raiz, que é a pasta compartilhada que a empresa usa no
    #drive pra salvar as fichas (Tive que usar o Drive Desktop). Aí dentro tem pastas de a até z, e dentro de cada uma
    #delas tem pastas nomeadas com as razões sociais dos clientes. Meu código vai verificar no caminho raiz se existe
    #pasta com a letra inicial da razão social do cliente que estamos gerando a ficha, se houver, ele entra na pasta, se nao,
    #ele cria ela e continua. Depois ele verifica se existe a pasta com o nome da razão social do cliente, se houver, ele entra,
    #se não, ele cria. Por fim ele cria a planilha dentro dessa pasta.
    caminho_base = f"G:\\Drives compartilhados\\FICHAS DE IMPLANTACAO"

    primeira_letra = dados_filtrados["Razão Social"][0].upper()
    subpasta_letra = os.path.join(caminho_base, primeira_letra)

    #cria a pasta da letra inicial da razao social do cliente
    tipo_mensagem_letra = "já existia" if os.path.exists(subpasta_letra) else "foi criada"
    os.makedirs(subpasta_letra, exist_ok=True)
    print(f"\nPasta da letra '{primeira_letra}'{tipo_mensagem_letra}: {subpasta_letra}")

    #cria a pasta da razão social do cliente
    pasta_razao_social = os.path.join(subpasta_letra, dados_filtrados["Razão Social"])
    tipo_mensagem_razao = "já existia" if os.path.exists(pasta_razao_social) else "foi criada"
    os.makedirs(pasta_razao_social, exist_ok=True)
    print(f"Pasta da Razão Social '{dados_filtrados['Razão Social']}' {tipo_mensagem_razao}: {pasta_razao_social}")

    #cria a planilha dentro da pasta da razão social do cliente
    arquivo_excel = os.path.join(pasta_razao_social, f"{dados_filtrados['Loja']}.xlsx")

    #essa parte faz um get na url onde está hospedada o logo da empresa que eu coloco dentro da planilha.
    caminho_imagem = f"G:\\Drives compartilhados\\FICHAS DE IMPLANTACAO\\payer.png"
    #aí finalmente eu chamo a função que cria a planilha.
    gerar_planilha_estilizada(dados_filtrados_reorganizado, arquivo_excel, caminho_imagem)
    os.startfile(pasta_razao_social)

    #isso aqui mantem o programa rodando por 3 seg e depois inicia o laço lá de cima de novo.
    #Aí ele apaga os arquivos gerados no processo para nao ficar resíduo.
    time.sleep(1)
    os.remove('teste') 
    os.remove('arquivo.json') 
